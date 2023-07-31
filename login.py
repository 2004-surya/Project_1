import openpyxl

def add_Employee_data():
    Employees = []

    while True:
        print("Enter Employee details:")
        name = input("Name: ")
        roll_no = input("Id no: ")
        branch = input("Company: ")
        phone_no = input("Phone No: ")

        Employee = {
            "Name": name,
            "Roll No": roll_no,
            "Branch": branch,
            "Phone No": phone_no,
        }
        Employees.append(Employee)

        more_data = input("Do you want to enter data for another employee? (yes/no): ")
        if more_data.lower() != "yes":
            break

    return Employees

def create_excel(Employees, output_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write headers
    headers = list(Employees[0].keys())
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

    # Write student data
    for row_idx, Employee in enumerate(Employees, start=2):
        for col_idx, value in enumerate(Employee.values(), start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    # Save the Excel file
    workbook.save(output_file)

if __name__ == "__main__":
    print("Enter student data. Type 'done' when you finish.")

    Employee_data = add_Employee_data()

    if Employee_data:
        print("\nEmployees data collected successfully.")
        output_file_name = input("Enter the output Excel file name: ")

        create_excel(Employee_data, output_file_name)
        print(f"Data successfully converted and saved to {output_file_name}.")
    else:
        print("No Employeech data to convert.")